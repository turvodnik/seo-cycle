#!/usr/bin/env python3
"""Ingest WriterZen browser exports as bounded keyword/source-pack evidence."""

from __future__ import annotations

import argparse
import csv
import io
import json
import math
import pathlib
import re
import sys
from collections import Counter, defaultdict
from typing import Any

from seo_cycle_core.config import find_config, load_yaml, nested_get, project_root_for, rel_path
from seo_cycle_core.source_artifacts import (
    compact_text,
    make_vector_record,
    read_cached_distillate,
    source_artifact_paths,
    stable_cache_key,
    utc_now_iso,
    write_source_artifacts,
)


PROVIDER = "writerzen"
DEFAULT_IMPORT_DIR = "seo/research/writerzen/imports"
MODES = ("auto", "topic_discovery", "keyword_explorer", "keyword_planner", "domain_focus", "plagiarism_export")

COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "keyword": ("keyword", "query", "phrase", "term", "topic", "search term", "ключ", "запрос", "тема"),
    "cluster": ("cluster", "group", "keyword group", "topic cluster", "parent topic", "topic", "кластер", "группа"),
    "volume": ("volume", "search volume", "vol", "avg. monthly searches", "monthly volume", "частотность", "показы", "спрос"),
    "cpc": ("cpc", "cost per click", "bid", "ставка", "цена клика"),
    "difficulty": ("difficulty", "kd", "keyword difficulty", "seo difficulty", "competition", "конкуренция", "сложность"),
    "allintitle": ("allintitle", "all in title", "allintitle count", "intitle", "в title"),
    "kgr": ("kgr", "golden filter", "golden score", "keyword golden ratio"),
    "intent": ("intent", "search intent", "интент", "намерение"),
    "buying_journey": ("buying journey", "journey", "funnel", "stage", "intent stage", "этап"),
    "serp_type": ("serp type", "serp feature", "serp features", "page type", "тип выдачи", "тип serp"),
    "brand": ("brand/non-brand", "brand", "branded", "brand type", "бренд"),
    "trend": ("trend", "trends", "seasonality", "динамика", "тренд"),
    "da": ("da", "domain authority", "domain rating", "dr", "authority"),
    "domain": ("domain", "competitor domain", "website", "site", "домен", "сайт"),
    "url": ("url", "page", "landing page", "адрес", "страница"),
}


def normalized_header(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip().lower().replace("_", " ").replace("-", " "))


def detect_delimiter(text: str) -> str:
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        return dialect.delimiter
    except csv.Error:
        return "\t" if "\t" in sample else ","


def read_csv_rows(path: pathlib.Path, *, max_rows: int) -> list[dict[str, Any]]:
    text = path.read_text(encoding="utf-8-sig", errors="replace")
    delimiter = detect_delimiter(text)
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    rows: list[dict[str, Any]] = []
    for idx, row in enumerate(reader):
        if idx >= max_rows:
            break
        rows.append({str(key or "").strip(): value for key, value in row.items()})
    return rows


def read_json_rows(path: pathlib.Path, *, max_rows: int) -> list[dict[str, Any]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    candidates: Any
    if isinstance(data, list):
        candidates = data
    elif isinstance(data, dict):
        candidates = data.get("rows") or data.get("items") or data.get("keywords") or data.get("data") or [data]
    else:
        candidates = []
    rows: list[dict[str, Any]] = []
    for item in candidates[:max_rows] if isinstance(candidates, list) else []:
        if isinstance(item, dict):
            rows.append(item)
        elif isinstance(item, str):
            rows.append({"keyword": item})
    return rows


def read_xlsx_rows(path: pathlib.Path, *, max_rows: int) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:  # pragma: no cover - optional dependency
        raise RuntimeError("XLSX import requires openpyxl. Export CSV from WriterZen or install openpyxl.") from exc
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(rows_iter, [])]
    rows: list[dict[str, Any]] = []
    for idx, values in enumerate(rows_iter):
        if idx >= max_rows:
            break
        rows.append({headers[i]: values[i] if i < len(values) else "" for i in range(len(headers))})
    return rows


def read_markdown_rows(path: pathlib.Path, *, max_rows: int) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for raw in path.read_text(encoding="utf-8", errors="replace").splitlines():
        line = raw.strip().strip("|")
        if not line or line.startswith("#") or set(line.replace("|", "").replace("-", "").strip()) == set():
            continue
        cells = [cell.strip() for cell in line.split("|")]
        if len(cells) == 1:
            rows.append({"keyword": cells[0]})
        elif len(cells) >= 2:
            rows.append({"keyword": cells[0], "note": cells[1]})
        if len(rows) >= max_rows:
            break
    return rows


def read_export(path: pathlib.Path, *, max_rows: int) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix in {".csv", ".tsv"}:
        return read_csv_rows(path, max_rows=max_rows)
    if suffix == ".json":
        return read_json_rows(path, max_rows=max_rows)
    if suffix == ".xlsx":
        return read_xlsx_rows(path, max_rows=max_rows)
    if suffix in {".md", ".txt"}:
        return read_markdown_rows(path, max_rows=max_rows)
    raise RuntimeError(f"Unsupported WriterZen export extension: {path.suffix}")


def value_for(row: dict[str, Any], field: str) -> Any:
    lowered = {normalized_header(str(key)): value for key, value in row.items()}
    for alias in COLUMN_ALIASES[field]:
        value = lowered.get(normalized_header(alias))
        if value not in (None, ""):
            return value
    return None


def parse_number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    text = str(value).strip().lower()
    multiplier = 1.0
    if text.endswith("k"):
        multiplier = 1000.0
        text = text[:-1]
    if text.endswith("m"):
        multiplier = 1_000_000.0
        text = text[:-1]
    text = re.sub(r"[^0-9.,-]", "", text).replace(",", ".")
    if not text or text in {"-", "."}:
        return None
    try:
        return float(text) * multiplier
    except ValueError:
        return None


def infer_mode(path: pathlib.Path, requested: str) -> str:
    if requested != "auto":
        return requested
    name = path.stem.lower()
    if "topic" in name:
        return "topic_discovery"
    if "planner" in name or "cluster" in name:
        return "keyword_planner"
    if "domain" in name or "competitor" in name:
        return "domain_focus"
    if "plag" in name:
        return "plagiarism_export"
    return "keyword_explorer"


def priority_score(row: dict[str, Any]) -> float:
    volume = float(row.get("volume") or 0)
    difficulty = row.get("difficulty")
    difficulty_score = 20.0
    if isinstance(difficulty, (int, float)):
        difficulty_score = max(0.0, min(40.0, (100.0 - float(difficulty)) * 0.4))
    intent = str(row.get("intent") or "").lower()
    buying = str(row.get("buying_journey") or "").lower()
    kgr = row.get("kgr")
    commercial_bonus = 10.0 if any(token in intent + " " + buying for token in ("buy", "commercial", "transaction", "purchase", "куп", "заказ", "цена")) else 0.0
    kgr_bonus = 8.0 if isinstance(kgr, (int, float)) and kgr <= 0.25 else 0.0
    volume_score = min(40.0, math.log10(max(volume, 0) + 1) * 12.0)
    return round(min(100.0, volume_score + difficulty_score + commercial_bonus + kgr_bonus), 2)


def normalize_row(row: dict[str, Any], *, source_file: str, mode: str) -> dict[str, Any] | None:
    keyword = value_for(row, "keyword")
    if keyword is None:
        return None
    normalized: dict[str, Any] = {
        "keyword": str(keyword).strip(),
        "source_file": source_file,
        "source_mode": mode,
        "provider": PROVIDER,
    }
    for field in ("cluster", "intent", "buying_journey", "serp_type", "brand", "trend", "domain", "url"):
        value = value_for(row, field)
        if value not in (None, ""):
            normalized[field] = str(value).strip()
    for field in ("volume", "cpc", "difficulty", "allintitle", "kgr", "da"):
        value = parse_number(value_for(row, field))
        if value is not None:
            normalized[field] = value
    normalized["priority_score"] = priority_score(normalized)
    return normalized


def collect_files(project_root: pathlib.Path, args: argparse.Namespace, cfg: dict[str, Any]) -> list[pathlib.Path]:
    files = [pathlib.Path(raw).expanduser() for raw in args.export_file]
    if files:
        return [path.resolve() for path in files]
    provider_cfg = cfg.get("writerzen_provider", {}) if isinstance(cfg.get("writerzen_provider"), dict) else {}
    import_dir = rel_path(project_root, args.import_dir or provider_cfg.get("import_dir") or DEFAULT_IMPORT_DIR)
    if not import_dir.exists():
        return []
    suffixes = {".csv", ".tsv", ".json", ".xlsx", ".md", ".txt"}
    return sorted(path for path in import_dir.iterdir() if path.is_file() and path.suffix.lower() in suffixes)


def summarize_rows(rows: list[dict[str, Any]], *, max_rows: int) -> dict[str, Any]:
    top = sorted(rows, key=lambda row: (float(row.get("priority_score") or 0), float(row.get("volume") or 0)), reverse=True)[:max_rows]
    counters: dict[str, Counter[str]] = {}
    for field in ("source_mode", "intent", "buying_journey", "serp_type", "cluster", "domain"):
        counter: Counter[str] = Counter()
        for row in rows:
            value = str(row.get(field) or "").strip()
            if value:
                counter[value] += 1
        counters[field] = counter
    opportunities = [
        row
        for row in top
        if float(row.get("priority_score") or 0) >= 55
        or str(row.get("intent") or "").lower() in {"commercial", "transactional", "buy"}
    ][:20]
    return {
        "row_count": len(rows),
        "top_keywords": top,
        "opportunities": opportunities,
        "counts": {field: dict(counter.most_common(20)) for field, counter in counters.items()},
    }


def render_markdown(distillate: dict[str, Any]) -> str:
    lines = [
        "# WriterZen Source Pack",
        "",
        f"- Topic: {distillate['topic']}",
        f"- Region: {distillate['region']}",
        f"- Status: `{distillate['status']}`",
        f"- Source type: `{distillate['source_type']}`",
        f"- Cache key: `{distillate['cache_key']}`",
        f"- Rows: {distillate['summary']['row_count']}",
        "",
        "## Summary",
        distillate.get("summary_text") or "",
        "",
        "## Top Keywords",
        "| Keyword | Volume | KD | Intent | Journey | SERP Type | Priority |",
        "|---|---:|---:|---|---|---|---:|",
    ]
    for row in distillate["summary"].get("top_keywords", [])[:20]:
        lines.append(
            "| {keyword} | {volume} | {difficulty} | {intent} | {buying_journey} | {serp_type} | {priority_score} |".format(
                keyword=str(row.get("keyword", "")).replace("|", "\\|"),
                volume=row.get("volume", ""),
                difficulty=row.get("difficulty", ""),
                intent=str(row.get("intent", "")).replace("|", "\\|"),
                buying_journey=str(row.get("buying_journey", "")).replace("|", "\\|"),
                serp_type=str(row.get("serp_type", "")).replace("|", "\\|"),
                priority_score=row.get("priority_score", ""),
            )
        )
    lines.extend(["", "## Intent / SERP / Cluster Counts"])
    for field in ("intent", "buying_journey", "serp_type", "cluster", "source_mode"):
        counts = distillate["summary"]["counts"].get(field, {})
        lines.append(f"### {field}")
        if counts:
            lines.extend(f"- {name}: {count}" for name, count in list(counts.items())[:12])
        else:
            lines.append("- none")
    lines.extend(
        [
            "",
            "## Browser Workflow",
            "1. Open WriterZen in the already logged-in browser session.",
            "2. Run Topic Discovery for broad topical expansion and export CSV/XLSX.",
            "3. Run Keyword Explorer for volume/CPC/Allintitle/KGR/trends and export CSV/XLSX.",
            "4. Run Keyword Planner for intent, Buying Journey, SERP Type, clusters, brand/non-brand and export CSV/XLSX.",
            "5. Put exports into `seo/research/writerzen/imports/` or pass them with `--export-file`.",
            "6. Re-run this source-pack; use only this distillate downstream, not the raw exports.",
            "",
            "## Source Policy",
            distillate.get("source_policy", "Use WriterZen distillate only downstream."),
        ]
    )
    return "\n".join(lines) + "\n"


def build_report(cfg_path: pathlib.Path, args: argparse.Namespace) -> dict[str, Any]:
    cfg = load_yaml(cfg_path)
    project_root = project_root_for(cfg_path)
    region = args.region or nested_get(cfg, "locale.country") or nested_get(cfg, "locale.region") or "global"
    topic = args.topic or nested_get(cfg, "project.name") or nested_get(cfg, "project.domain") or "writerzen"
    files = collect_files(project_root, args, cfg)

    all_rows: list[dict[str, Any]] = []
    raw_files: list[dict[str, Any]] = []
    errors: list[str] = []
    for path in files:
        mode = infer_mode(path, args.mode)
        try:
            exported_rows = read_export(path, max_rows=args.max_input_rows)
        except Exception as exc:
            errors.append(f"{path.name}: {exc}")
            continue
        normalized_rows = [
            normalized
            for row in exported_rows
            if (normalized := normalize_row(row, source_file=path.name, mode=mode)) is not None
        ]
        all_rows.extend(normalized_rows)
        raw_files.append(
            {
                "file": str(path),
                "mode": mode,
                "rows_loaded": len(exported_rows),
                "rows_normalized": len(normalized_rows),
            }
        )

    source_type = "browser_export" if all_rows else "unavailable"
    cache_key = stable_cache_key(
        {
            "provider": PROVIDER,
            "topic": topic,
            "region": region,
            "mode": args.mode,
            "files": raw_files,
            "rows": all_rows,
        },
        label=str(topic),
    )

    if all_rows:
        status = "ready"
        summary = summarize_rows(all_rows, max_rows=args.max_distillate_rows)
        summary_text = (
            f"WriterZen browser exports produced {len(all_rows)} normalized rows from {len(raw_files)} file(s). "
            "Use Topic Discovery for breadth, Keyword Explorer for Google metrics, and Keyword Planner for intent/clusters."
        )
    else:
        status = "fallback_required"
        summary = {"row_count": 0, "top_keywords": [], "opportunities": [], "counts": {}}
        summary_text = "No usable WriterZen export rows were supplied. Open WriterZen in a logged-in browser, export CSV/XLSX, then rerun."

    raw_payload = {
        "provider": PROVIDER,
        "status": status,
        "topic": topic,
        "region": region,
        "source_type": source_type,
        "created_at": utc_now_iso(),
        "raw_files": raw_files,
        "normalized_rows": all_rows,
        "errors": errors,
    }
    distillate = {
        "provider": PROVIDER,
        "status": status,
        "cache_key": cache_key,
        "topic": topic,
        "region": region,
        "source_type": source_type,
        "summary_text": compact_text(summary_text, max_chars=args.max_summary_chars),
        "summary": summary,
        "errors": errors,
        "source_policy": "WriterZen has no public API in this workflow. Keep raw browser exports on disk; downstream prompts use only bounded distillates/vector records.",
        "browser_workflow_required": True,
        "stores_password": False,
        "paid_api_used": False,
    }
    markdown = render_markdown(distillate)
    paths: dict[str, str] = {}
    if args.write:
        cached = read_cached_distillate(project_root, PROVIDER, cache_key)
        if cached:
            cached_paths = source_artifact_paths(project_root, PROVIDER, cache_key)
            return {
                "provider": PROVIDER,
                "status": "cache_hit",
                "generated_at": utc_now_iso(),
                "cache_key": cache_key,
                "topic": topic,
                "region": region,
                "source_type": source_type,
                "raw_files": raw_files,
                "errors": errors,
                "distillate": cached,
                "paths": {key: str(path) for key, path in cached_paths.items()},
                "writes_to_site": False,
                "stores_password": False,
                "paid_api_used": False,
            }
        paths = write_source_artifacts(
            project_root,
            PROVIDER,
            cache_key,
            raw_payload=raw_payload,
            distillate_markdown=markdown,
            distillate_payload=distillate,
            vector_record=make_vector_record(
                provider=PROVIDER,
                cache_key=cache_key,
                topic=str(topic),
                region=str(region),
                mode=source_type,
                status=status,
                summary=summary_text[:1000],
                citations=[],
                metadata={
                    "source_type": source_type,
                    "files": [row["file"] for row in raw_files],
                    "rows": len(all_rows),
                    "modes": sorted({row.get("mode", "") for row in raw_files}),
                    "top_keywords": [row.get("keyword") for row in summary.get("top_keywords", [])[:10]],
                },
            ),
        )
    return {
        "provider": PROVIDER,
        "status": status,
        "generated_at": utc_now_iso(),
        "cache_key": cache_key,
        "topic": topic,
        "region": region,
        "source_type": source_type,
        "raw_files": raw_files,
        "errors": errors,
        "distillate": distillate,
        "paths": paths,
        "writes_to_site": False,
        "stores_password": False,
        "paid_api_used": False,
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("config", nargs="?", help="Path to seo-cycle.yaml")
    parser.add_argument("--topic", help="Topic/seed represented by this WriterZen export.")
    parser.add_argument("--region", help="Target region label.")
    parser.add_argument("--mode", choices=MODES, default="auto", help="WriterZen export type.")
    parser.add_argument("--export-file", action="append", default=[], help="WriterZen CSV/TSV/JSON/XLSX/MD export file. Repeatable.")
    parser.add_argument("--import-dir", help=f"Directory with WriterZen exports, default {DEFAULT_IMPORT_DIR}.")
    parser.add_argument("--max-input-rows", type=int, default=5000)
    parser.add_argument("--max-distillate-rows", type=int, default=40)
    parser.add_argument("--max-summary-chars", type=int, default=3000)
    parser.add_argument("--write", action="store_true")
    parser.add_argument("--format", choices=("md", "json"), default="md")
    args = parser.parse_args()

    cfg_path = pathlib.Path(args.config).expanduser().resolve() if args.config else find_config(pathlib.Path.cwd())
    if not cfg_path or not cfg_path.exists():
        print(f"ERROR: seo-cycle.yaml not found in {pathlib.Path.cwd()}", file=sys.stderr)
        return 2
    try:
        report = build_report(cfg_path, args)
    except Exception as exc:  # pragma: no cover - CLI boundary
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    if args.format == "json":
        print(json.dumps(report, ensure_ascii=False, indent=2, sort_keys=True))
    else:
        print(render_markdown(report["distillate"]), end="")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
